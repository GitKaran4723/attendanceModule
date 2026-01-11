"""
Fee Management Routes
=====================
API endpoints for managing student fees, receipts, and fee structures.

Routes:
- Student routes: View fees, add receipts
- Faculty routes: Set fee structure, manage receipts, approve
- Admin routes: Reports, defaulters, analytics
- Parent routes: View fees (read-only)
"""

from flask import Blueprint, request, jsonify, render_template, session, send_file
from datetime import datetime, date
from sqlalchemy import func, and_, or_
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from models import db, FeeStructure, FeeReceipt, Student, Section, User, Faculty
from auth import login_required, student_required, faculty_required, admin_required, get_current_user
from fee_permissions import (
    can_set_fee_structure, can_edit_receipt, can_approve_receipt,
    can_delete_receipt, can_view_student_fees, get_class_teacher_students,
    is_class_teacher_of_student, get_class_teacher_sections
)

# Create blueprint
fee_bp = Blueprint('fees', __name__, url_prefix='/fees')


# =============================================================================
# STUDENT ROUTES
# =============================================================================

@fee_bp.route('/student/fees', methods=['GET'])
@student_required
def student_fees():
    """Student fee dashboard - view own fee structures and receipts"""
    current_user = get_current_user()
    student = current_user.student
    
    if not student:
        return jsonify({'error': 'Student profile not found'}), 404
    
    # Get all fee structures for this student
    fee_structures = FeeStructure.query.filter_by(
        student_id=student.student_id,
        is_deleted=False
    ).order_by(FeeStructure.academic_year.desc()).all()
    
    # Format data for response
    fee_data = []
    for structure in fee_structures:
        receipts = structure.fee_receipts.filter_by(is_deleted=False).all()
        fee_data.append({
            'structure': structure.to_dict(),
            'receipts': [receipt.to_dict() for receipt in receipts],
            'pending_receipts': len([r for r in receipts if not r.approved]),
            'approved_receipts': len([r for r in receipts if r.approved])
        })
    
    return jsonify({
        'success': True,
        'student': {
            'student_id': student.student_id,
            'name': student.name,
            'roll_number': student.roll_number,
            'section': student.section.section_name if student.section else None,
            'joining_year': student.joining_academic_year
        },
        'fee_structures': fee_data
    })


@fee_bp.route('/student/add-receipt', methods=['POST'])
@student_required
def student_add_receipt():
    """Student adds a new fee receipt"""
    current_user = get_current_user()
    student = current_user.student
    
    if not student:
        return jsonify({'error': 'Student profile not found'}), 404
    
    data = request.get_json()
    
    # Validate required fields
    required_fields = ['fee_structure_id', 'receipt_number', 'receipt_phone', 
                      'amount_paid', 'payment_date']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    # Check if fee structure exists and belongs to student
    fee_structure = FeeStructure.query.get(data['fee_structure_id'])
    if not fee_structure or fee_structure.student_id != student.student_id:
        return jsonify({'error': 'Invalid fee structure'}), 404
    
    # Check receipt number uniqueness
    existing_receipt = FeeReceipt.query.filter_by(receipt_number=data['receipt_number']).first()
    if existing_receipt:
        return jsonify({'error': 'Receipt number already exists'}), 400
    
    # Validate amount
    try:
        amount = float(data['amount_paid'])
        if amount <= 0:
            return jsonify({'error': 'Amount must be greater than 0'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid amount'}), 400
    
    # Parse payment date
    try:
        payment_date = datetime.strptime(data['payment_date'], '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    
    # Create receipt
    receipt = FeeReceipt(
        fee_structure_id=fee_structure.fee_structure_id,
        student_id=student.student_id,
        receipt_number=data['receipt_number'],
        receipt_phone=data['receipt_phone'],
        amount_paid=amount,
        payment_date=payment_date,
        payment_mode=data.get('payment_mode', 'cash'),
        entered_by_user_id=current_user.user_id,
        entered_by_role='student',
        approved=False,  # Pending approval
        remarks=data.get('remarks', '')
    )
    
    db.session.add(receipt)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Receipt added successfully. Waiting for class teacher approval.',
        'receipt': receipt.to_dict()
    }), 201


@fee_bp.route('/student/summary', methods=['GET'])
@student_required
def student_fee_summary():
    """Get fee summary for student dashboard"""
    current_user = get_current_user()
    student = current_user.student
    
    if not student:
        return jsonify({'error': 'Student profile not found'}), 404
    
    # Get current academic year fee structure
    current_year = datetime.now().year
    current_academic_year = f"{current_year}-{current_year + 1}"
    
    current_structure = FeeStructure.query.filter_by(
        student_id=student.student_id,
        academic_year=current_academic_year,
        is_deleted=False
    ).first()
    
    summary = {
        'current_year': current_academic_year,
        'has_fee_structure': current_structure is not None
    }
    
    if current_structure:
        pending_receipts = current_structure.fee_receipts.filter_by(approved=False, is_deleted=False).count()
        summary.update({
            'total_fees': current_structure.total_fees,
            'total_paid': current_structure.calculate_total_paid(),
            'outstanding': current_structure.calculate_outstanding(),
            'pending_approvals': pending_receipts
        })
    else:
        summary.update({
            'total_fees': 0,
            'total_paid': 0,
            'outstanding': 0,
            'pending_approvals': 0,
            'message': 'Fees not yet assigned for this year'
        })
    
    return jsonify({'success': True, 'summary': summary})


# =============================================================================
# FACULTY/CLASS TEACHER ROUTES
# =============================================================================

@fee_bp.route('/faculty/students', methods=['GET'])
@faculty_required
def faculty_students_fees():
    """Faculty view fee records for students in assigned sections"""
    current_user = get_current_user()
    faculty = current_user.faculty
    
    if not faculty:
        return jsonify({'error': 'Faculty profile not found'}), 404
    
    # Get sections where faculty is class teacher
    sections = get_class_teacher_sections(faculty.faculty_id)
    
    if not sections:
        return jsonify({
            'success': True,
            'message': 'You are not assigned as class teacher for any section',
            'sections': [],
            'students': []
        })
    
    # Get all students in these sections
    student_ids = get_class_teacher_students(faculty.faculty_id)
    students = Student.query.filter(
        Student.student_id.in_(student_ids),
        Student.is_deleted == False
    ).all()
    
    # Filters from query params
    section_filter = request.args.get('section_id')
    approval_status = request.args.get('approval_status')  # 'pending', 'approved', 'all'
    
    # Format data with sorting metadata
    students_data = []
    for student in students:
        # Skip if section filter doesn't match
        if section_filter and student.section_id != section_filter:
            continue
        
        # Strictly use student's current academic year as per user request
        academic_year = student.current_academic_year or '2025-2026'
        
        # Get fee structure for this student and year (DO NOT AUTO-CREATE)
        structure = FeeStructure.query.filter_by(
            student_id=student.student_id,
            academic_year=academic_year,
            is_deleted=False
        ).first()
        
        # Determine sorting priority
        # Priority 1: No fee structure (highest priority)
        # Priority 2: Has outstanding fees
        # Priority 3: Fully paid (lowest priority)
        
        if not structure:
            # No fee structure set - highest priority
            sort_priority = 1
            outstanding = 0
            total_fees = 0
            total_paid = 0
            receipts = []
            
            # Create minimal structure dict for display
            structure_dict = {
                'fee_structure_id': None,
                'student_id': student.student_id,
                'student_name': student.name,
                'roll_number': student.roll_number or 'N/A',
                'section_name': student.section.section_name if student.section else 'N/A',
                'joining_year': student.joining_academic_year,
                'academic_year': academic_year,
                'total_fees': 0,
                'total_paid': 0,
                'outstanding': 0,
                'base_fees': 0,
                'carry_forward': 0,
                'additional_charges': 0,
                'status': 'not_set',
                'message': 'Fee structure not set by admin'
            }
        else:
            # Fee structure exists
            total_fees = structure.total_fees
            total_paid = structure.calculate_total_paid()
            outstanding = structure.calculate_outstanding()
            
            # Get receipts for this structure
            receipts_query = FeeReceipt.query.filter_by(
                fee_structure_id=structure.fee_structure_id,
                is_deleted=False
            )
            
            if approval_status == 'pending':
                receipts_query = receipts_query.filter_by(approved=False)
            elif approval_status == 'approved':
                receipts_query = receipts_query.filter_by(approved=True)
            
            receipts = receipts_query.all()
            
            # Determine priority based on payment status
            if outstanding > 0:
                sort_priority = 2  # Has outstanding fees
            else:
                sort_priority = 3  # Fully paid
            
            # Add student info to structure dict
            structure_dict = structure.to_dict()
            structure_dict['student_name'] = student.name
            structure_dict['roll_number'] = student.roll_number or 'N/A'
            structure_dict['section_name'] = student.section.section_name if student.section else 'N/A'
            structure_dict['joining_year'] = student.joining_academic_year
        
        students_data.append({
            'structure': structure_dict,
            'receipts': [r.to_dict() for r in receipts],
            'sort_priority': sort_priority,
            'outstanding': outstanding,
            'student_name': student.name  # For secondary sorting
        })
    
    # Sort students:
    # 1. By priority (1=no structure, 2=outstanding, 3=fully paid)
    # 2. Within same priority, sort by name alphabetically
    students_data.sort(key=lambda x: (x['sort_priority'], x['student_name'].lower()))
    
    # Remove sorting metadata before returning
    for item in students_data:
        del item['sort_priority']
        del item['outstanding']
        del item['student_name']
    
    # Calculate summary stats for the dashboard
    pending_count = sum(1 for item in students_data if item['structure'].get('outstanding', 0) > 0)
    
    return jsonify({
        'success': True,
        'sections': [{'section_id': s.section_id, 'section_name': s.section_name} for s in sections],
        'summary_stats': {
            'pending_payment_count': pending_count,
            'total_students': len(students_data)
        },
        'data': students_data
    })


@fee_bp.route('/faculty/add-receipt', methods=['POST'])
@faculty_required
def faculty_add_receipt():
    """Faculty adds receipt for a student"""
    current_user = get_current_user()
    faculty = current_user.faculty
    
    if not faculty:
        return jsonify({'error': 'Faculty profile not found'}), 404
    
    data = request.get_json()
    
    # Validate required fields
    required_fields = ['fee_structure_id', 'receipt_number', 'receipt_phone', 
                      'amount_paid', 'payment_date']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    # Get fee structure
    fee_structure = FeeStructure.query.get(data['fee_structure_id'])
    if not fee_structure:
        return jsonify({'error': 'Fee structure not found'}), 404
    
    # Check permission (must be class teacher of student)
    if not is_class_teacher_of_student(faculty.faculty_id, fee_structure.student_id):
        return jsonify({'error': 'You can only add receipts for students in your sections'}), 403
    
    # Check receipt number uniqueness
    existing_receipt = FeeReceipt.query.filter_by(receipt_number=data['receipt_number']).first()
    if existing_receipt:
        return jsonify({'error': 'Receipt number already exists'}), 400
    
    # Validate amount
    try:
        amount = float(data['amount_paid'])
        if amount <= 0:
            return jsonify({'error': 'Amount must be greater than 0'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid amount'}), 400
    
    # Parse payment date
    try:
        payment_date = datetime.strptime(data['payment_date'], '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    
    # Create receipt
    approve_immediately = data.get('approve_immediately', False)
    
    receipt = FeeReceipt(
        fee_structure_id=fee_structure.fee_structure_id,
        student_id=fee_structure.student_id,
        receipt_number=data['receipt_number'],
        receipt_phone=data['receipt_phone'],
        amount_paid=amount,
        payment_date=payment_date,
        payment_mode=data.get('payment_mode', 'cash'),
        entered_by_user_id=current_user.user_id,
        entered_by_role='faculty',
        approved=approve_immediately,
        approved_by_user_id=current_user.user_id if approve_immediately else None,
        approved_at=datetime.utcnow() if approve_immediately else None,
        remarks=data.get('remarks', '')
    )
    
    db.session.add(receipt)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Receipt added successfully' + (' and approved' if approve_immediately else ''),
        'receipt': receipt.to_dict()
    }), 201


@fee_bp.route('/faculty/receipt/<receipt_id>/edit', methods=['PUT'])
@faculty_required
def faculty_edit_receipt(receipt_id):
    """Faculty edits a receipt (can edit even approved receipts)"""
    current_user = get_current_user()
    faculty = current_user.faculty
    
    if not faculty:
        return jsonify({'error': 'Faculty profile not found'}), 404
    
    receipt = FeeReceipt.query.get(receipt_id)
    if not receipt or receipt.is_deleted:
        return jsonify({'error': 'Receipt not found'}), 404
    
    # Check permission
    if not can_edit_receipt(receipt, current_user):
        return jsonify({'error': 'You do not have permission to edit this receipt'}), 403
    
    data = request.get_json()
    
    # Update fields
    if 'receipt_number' in data and data['receipt_number'] != receipt.receipt_number:
        # Check uniqueness if changing receipt number
        existing = FeeReceipt.query.filter_by(receipt_number=data['receipt_number']).first()
        if existing:
            return jsonify({'error': 'Receipt number already exists'}), 400
        receipt.receipt_number = data['receipt_number']
    
    if 'receipt_phone' in data:
        receipt.receipt_phone = data['receipt_phone']
    
    if 'amount_paid' in data:
        try:
            receipt.amount_paid = float(data['amount_paid'])
        except ValueError:
            return jsonify({'error': 'Invalid amount'}), 400
    
    if 'payment_date' in data:
        try:
            receipt.payment_date = datetime.strptime(data['payment_date'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format'}), 400
    
    if 'payment_mode' in data:
        receipt.payment_mode = data['payment_mode']
    
    if 'remarks' in data:
        receipt.remarks = data['remarks']
    
    receipt.updated_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Receipt updated successfully',
        'receipt': receipt.to_dict()
    })


@fee_bp.route('/faculty/receipt/<receipt_id>/approve', methods=['POST'])
@faculty_required
def faculty_approve_receipt(receipt_id):
    """Faculty approves a fee receipt"""
    current_user = get_current_user()
    faculty = current_user.faculty
    
    if not faculty:
        return jsonify({'error': 'Faculty profile not found'}), 404
    
    receipt = FeeReceipt.query.get(receipt_id)
    if not receipt or receipt.is_deleted:
        return jsonify({'error': 'Receipt not found'}), 404
    
    # Check permission
    if not can_approve_receipt(receipt, current_user):
        return jsonify({'error': 'You do not have permission to approve this receipt'}), 403
    
    if receipt.approved:
        return jsonify({'message': 'Receipt is already approved'})
    
    # Approve receipt
    receipt.approved = True
    receipt.approved_by_user_id = current_user.user_id
    receipt.approved_at = datetime.utcnow()
    receipt.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Receipt approved successfully',
        'receipt': receipt.to_dict()
    })


@fee_bp.route('/faculty/receipt/<receipt_id>', methods=['DELETE'])
@faculty_required
def faculty_delete_receipt(receipt_id):
    """Faculty deletes a receipt (soft delete)"""
    current_user = get_current_user()
    faculty = current_user.faculty
    
    if not faculty:
        return jsonify({'error': 'Faculty profile not found'}), 404
    
    receipt = FeeReceipt.query.get(receipt_id)
    if not receipt or receipt.is_deleted:
        return jsonify({'error': 'Receipt not found'}), 404
    
    # Check permission
    if not can_delete_receipt(receipt, current_user):
        return jsonify({'error': 'You do not have permission to delete this receipt'}), 403
    
    # Soft delete
    receipt.is_deleted = True
    receipt.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Receipt deleted successfully'
    })


# =============================================================================
# ADMIN ROUTES
# =============================================================================

@fee_bp.route('/admin/structures', methods=['GET'])
@admin_required
def admin_fee_structures():
    """Admin view all fee structures with filters"""
    # Filters
    program_id = request.args.get('program_id')
    section_id = request.args.get('section_id')
    academic_year = request.args.get('academic_year')
    
    query = FeeStructure.query.filter_by(is_deleted=False)
    
    if section_id:
        query = query.filter_by(section_id=section_id)
    if academic_year:
        query = query.filter_by(academic_year=academic_year)
    
    structures = query.all()
    
    return jsonify({
        'success': True,
        'count': len(structures),
        'structures': [s.to_dict() for s in structures]
    })


@fee_bp.route('/admin/receipts', methods=['GET'])
@admin_required
def admin_fee_receipts():
    """Admin view all receipts with filters"""
    # Filters
    section_id = request.args.get('section_id')
    academic_year = request.args.get('academic_year')
    approved = request.args.get('approved')  # 'true', 'false', or None for all
    
    query = FeeReceipt.query.filter_by(is_deleted=False)
    
    if approved == 'true':
        query = query.filter_by(approved=True)
    elif approved == 'false':
        query = query.filter_by(approved=False)
    
    receipts = query.all()
    
    return jsonify({
        'success': True,
        'count': len(receipts),
        'receipts': [r.to_dict() for r in receipts]
    })


@fee_bp.route('/admin/defaulters', methods=['GET'])
@admin_required
def admin_fee_defaulters():
    """
    Admin comprehensive defaulter report
    Shows students with outstanding fees grouped by section/class teacher
    """
    # Filters
    program_id = request.args.get('program_id')
    section_id = request.args.get('section_id')
    academic_year = request.args.get('academic_year')
    class_teacher_id = request.args.get('class_teacher_id')
    min_outstanding = request.args.get('min_outstanding', type=float)
    max_outstanding = request.args.get('max_outstanding', type=float)
    
    # Get all fee structures
    query = FeeStructure.query.filter_by(is_deleted=False)
    
    if section_id:
        query = query.filter_by(section_id=section_id)
    if academic_year:
        query = query.filter_by(academic_year=academic_year)
    
    structures = query.all()
    
    # Build defaulter data
    defaulters = []
    total_outstanding = 0
    section_summary = {}
    teacher_summary = {}
    
    for structure in structures:
        outstanding = structure.calculate_outstanding()
        
        # Apply filters
        if min_outstanding and outstanding < min_outstanding:
            continue
        if max_outstanding and outstanding > max_outstanding:
            continue
        
        # Only include if there's outstanding balance
        if outstanding <= 0:
            continue
        
        student = structure.student
        section = structure.section
        
        # Class teacher filter
        if class_teacher_id and section.class_teacher_id != class_teacher_id:
            continue
        
        # Get class teacher info
        class_teacher_name = None
        if section and section.class_teacher:
            class_teacher_name = f"{section.class_teacher.first_name} {section.class_teacher.last_name}"
        
        # Get last payment date
        last_receipt = structure.fee_receipts.filter_by(approved=True, is_deleted=False).order_by(
            FeeReceipt.payment_date.desc()
        ).first()
        
        defaulter_data = {
            'student_id': student.student_id,
            'student_name': student.name,
            'roll_number': student.roll_number,
            'section_id': section.section_id if section else None,
            'section_name': section.section_name if section else None,
            'class_teacher_id': section.class_teacher_id if section else None,
            'class_teacher_name': class_teacher_name,
            'academic_year': structure.academic_year,
            'total_fees': structure.total_fees,
            'total_paid': structure.calculate_total_paid(),
            'outstanding': outstanding,
            'last_payment_date': last_receipt.payment_date.isoformat() if last_receipt else None
        }
        
        defaulters.append(defaulter_data)
        total_outstanding += outstanding
        
        # Build section summary
        section_key = section.section_id if section else 'no_section'
        if section_key not in section_summary:
            section_summary[section_key] = {
                'section_name': section.section_name if section else 'No Section',
                'count': 0,
                'total_outstanding': 0
            }
        section_summary[section_key]['count'] += 1
        section_summary[section_key]['total_outstanding'] += outstanding
        
        # Build teacher summary
        if class_teacher_name:
            teacher_key = section.class_teacher_id
            if teacher_key not in teacher_summary:
                teacher_summary[teacher_key] = {
                    'teacher_name': class_teacher_name,
                    'sections': set(),
                    'count': 0,
                    'total_outstanding': 0
                }
            teacher_summary[teacher_key]['sections'].add(section.section_name)
            teacher_summary[teacher_key]['count'] += 1
            teacher_summary[teacher_key]['total_outstanding'] += outstanding
    
    # Convert sets to lists for JSON
    for teacher_data in teacher_summary.values():
        teacher_data['sections'] = list(teacher_data['sections'])
    
    return jsonify({
        'success': True,
        'total_defaulters': len(defaulters),
        'total_outstanding_amount': round(total_outstanding, 2),
        'defaulters': defaulters,
        'section_summary': list(section_summary.values()),
        'teacher_summary': list(teacher_summary.values())
    })


@fee_bp.route('/admin/defaulters/export', methods=['GET'])
@admin_required
def admin_export_defaulters():
    """Export defaulter report to Excel"""
    # Get filters from query params
    params = {k: v for k, v in request.args.items()}
    
    # Build query for defaulters (reuse logic)
    query = FeeStructure.query.filter_by(is_deleted=False)
    
    if 'section_id' in params:
        query = query.filter_by(section_id=params['section_id'])
    if 'academic_year' in params:
        query = query.filter_by(academic_year=params['academic_year'])
    
    structures = query.all()
    
    # Collect defaulter data
    defaulters = []
    section_summary = {}
    teacher_summary = {}
    
    for structure in structures:
        outstanding = structure.calculate_outstanding()
        if outstanding <= 0:
            continue
        
        student = structure.student
        section = structure.section
        
        class_teacher_name = None
        if section and section.class_teacher:
            class_teacher_name = f"{section.class_teacher.first_name} {section.class_teacher.last_name}"
        
        last_receipt = structure.fee_receipts.filter_by(approved=True, is_deleted=False).order_by(
            FeeReceipt.payment_date.desc()
        ).first()
        
        defaulters.append({
            'student_name': student.name,
            'roll_number': student.roll_number,
            'section_name': section.section_name if section else 'N/A',
            'class_teacher_name': class_teacher_name or 'N/A',
            'academic_year': structure.academic_year,
            'total_fees': structure.total_fees,
            'total_paid': structure.calculate_total_paid(),
            'outstanding': outstanding,
            'last_payment_date': last_receipt.payment_date.isoformat() if last_receipt else 'N/A'
        })
        
        # Section summary
        section_key = section.section_name if section else 'No Section'
        if section_key not in section_summary:
            section_summary[section_key] = {'count': 0, 'total': 0}
        section_summary[section_key]['count'] += 1
        section_summary[section_key]['total'] += outstanding
        
        # Teacher summary
        if class_teacher_name:
            if class_teacher_name not in teacher_summary:
                teacher_summary[class_teacher_name] = {'sections': set(), 'count': 0, 'total': 0}
            teacher_summary[class_teacher_name]['sections'].add(section.section_name)
            teacher_summary[class_teacher_name]['count'] += 1
            teacher_summary[class_teacher_name]['total'] += outstanding
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Detailed Report
    ws1 = wb.active
    ws1.title = "Defaulters Detail"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Student Name', 'Roll Number', 'Section', 'Class Teacher', 
               'Academic Year', 'Total Fees', 'Total Paid', 'Outstanding', 'Last Payment']
    ws1.append(headers)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for d in defaulters:
        ws1.append([
            d['student_name'], d['roll_number'], d['section_name'], d['class_teacher_name'],
            d['academic_year'], d['total_fees'], d['total_paid'], d['outstanding'], d['last_payment_date']
        ])
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Section Summary
    ws2 = wb.create_sheet("Section Summary")
    ws2.append(['Section', 'Student Count', 'Total Outstanding'])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for section_name, data in section_summary.items():
        ws2.append([section_name, data['count'], round(data['total'], 2)])
    
    # Sheet 3: Teacher Summary
    ws3 = wb.create_sheet("Teacher Summary")
    ws3.append(['Class Teacher', 'Sections', 'Student Count', 'Total Outstanding'])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for teacher_name, data in teacher_summary.items():
        ws3.append([teacher_name, ', '.join(data['sections']), data['count'], round(data['total'], 2)])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"Fee_Defaulters_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@fee_bp.route('/admin/reports', methods=['GET'])
@admin_required
def admin_fee_reports():
    """Admin fee collection reports and analytics"""
    academic_year = request.args.get('academic_year')
    
    # Get all structures for the year
    query = FeeStructure.query.filter_by(is_deleted=False)
    if academic_year:
        query = query.filter_by(academic_year=academic_year)
    
    structures = query.all()
    
    total_fees_assigned = sum(s.total_fees for s in structures)
    total_collected = sum(s.calculate_total_paid() for s in structures)
    total_outstanding = sum(s.calculate_outstanding() for s in structures)
    
    # Pending approvals count
    pending_approvals = FeeReceipt.query.filter_by(approved=False, is_deleted=False).count()
    
    return jsonify({
        'success': True,
        'analytics': {
            'total_students': len(structures),
            'total_fees_assigned': round(total_fees_assigned, 2),
            'total_collected': round(total_collected, 2),
            'total_outstanding': round(total_outstanding, 2),
            'collection_percentage': round((total_collected / total_fees_assigned * 100), 2) if total_fees_assigned > 0 else 0,
            'pending_approvals': pending_approvals
        }
    })


# =============================================================================
# PARENT ROUTES
# =============================================================================

@fee_bp.route('/parent/fees/<student_id>', methods=['GET'])
@login_required
def parent_view_fees(student_id):
    """Parent view student fee status (read-only)"""
    current_user = get_current_user()
    
    # Check permission
    if not can_view_student_fees(student_id, current_user):
        return jsonify({'error': 'You do not have permission to view this student\'s fees'}), 403
    
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'error': 'Student not found'}), 404
    
    # Get all fee structures for this student
    fee_structures = FeeStructure.query.filter_by(
        student_id=student_id,
        is_deleted=False
    ).order_by(FeeStructure.academic_year.desc()).all()
    
    # Format data
    fee_data = []
    for structure in fee_structures:
        receipts = structure.fee_receipts.filter_by(is_deleted=False, approved=True).all()
        fee_data.append({
            'structure': structure.to_dict(),
            'approved_receipts': [receipt.to_dict() for receipt in receipts]
        })
    
    return jsonify({
        'success': True,
        'student': {
            'student_id': student.student_id,
            'name': f"{student.name}",
            'roll_number': student.roll_number,
            'section': student.section.section_name if student.section else None
        },
        'fee_structures': fee_data
    })

